from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from datetime import datetime
from io import BytesIO
from typing import Optional

from app.core.database import get_db
from app.models.user import User, UserRole
from app.models.candidate import Candidate, CandidateStatus
from app.api.auth import get_current_user

router = APIRouter()


@router.get("/export")
async def export_candidates(
    status: Optional[CandidateStatus] = None,
    project_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """导出候选人数据为 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    query = select(Candidate).options(
        selectinload(Candidate.project),
        selectinload(Candidate.channel),
        selectinload(Candidate.owner)
    )
    
    # 数据权限
    if current_user.role == UserRole.RECRUITER:
        query = query.where(Candidate.owner_id == current_user.id)
    
    # 筛选条件
    if status:
        query = query.where(Candidate.status == status)
    if project_id:
        query = query.where(Candidate.project_id == project_id)
    
    query = query.order_by(Candidate.created_at.desc()).limit(1000)
    result = await db.execute(query)
    candidates = result.scalars().all()
    
    # 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "候选人列表"
    
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 表头
    headers = ["序号", "姓名", "手机号", "微信号", "年龄", "状态", "应聘项目", "渠道", "跟进人", "最后跟进", "录入时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 状态映射
    status_map = {
        "lead": "线索",
        "wechat": "已加微信",
        "interview": "已约面",
        "interviewed": "已到面",
        "onboarded": "已入职",
        "lost": "流失"
    }
    
    # 数据行
    for row, c in enumerate(candidates, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=c.name)
        ws.cell(row=row, column=3, value=c.phone)
        ws.cell(row=row, column=4, value=c.wechat or "-")
        ws.cell(row=row, column=5, value=c.age or "-")
        ws.cell(row=row, column=6, value=status_map.get(c.status.value, c.status.value))
        ws.cell(row=row, column=7, value=c.project.name if c.project else "-")
        ws.cell(row=row, column=8, value=c.channel.name if c.channel else "-")
        ws.cell(row=row, column=9, value=c.owner.real_name or c.owner.username if c.owner else "-")
        ws.cell(row=row, column=10, value=c.last_follow_at.strftime("%Y-%m-%d %H:%M") if c.last_follow_at else "-")
        ws.cell(row=row, column=11, value=c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "-")
    
    # 调整列宽
    column_widths = [6, 10, 15, 15, 6, 10, 20, 12, 10, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
